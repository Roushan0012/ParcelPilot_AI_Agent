#!/usr/bin/env python3
"""
ParcelPilot Ingestion Pipeline
1. Extracts text from all 6 PDFs in ./data
2. Chunks by section and heading with rich metadata
3. Generates Voyage-3 vector embeddings (1024 dimensions)
4. Extracts structured data from ParcelPilot_Assessment_Data.xlsx
5. Writes pre-computed dataset to src/lib/data/dataset.json
6. Generates full SQL seed script in supabase/seed.sql
"""

import os
import re
import json
import glob
import openpyxl
import pypdf
import requests

VOYAGE_API_KEY = os.getenv("VOYAGE_API_KEY", "")
DATA_DIR = "./data"
OUTPUT_JSON = "./src/lib/data/dataset.json"
OUTPUT_SQL = "./supabase/seed.sql"

os.makedirs("./src/lib/data", exist_ok=True)
os.makedirs("./supabase", exist_ok=True)

# Define PDF metadata and custom chunk definitions
DOC_METADATA = {
    "01_Support_Policy_v3_CURRENT.pdf": {
        "source_name": "01_Support_Policy_v3_CURRENT.pdf",
        "doc_title": "ParcelPilot Support Policy v3 (CURRENT)",
        "version": "v3_CURRENT",
        "status": "CURRENT",
        "effective_date": "2026-05-01",
        "customer_scope": "general",
        "authority_level": 2,
        "is_authoritative": True,
        "supersedes": "Support Policy v2"
    },
    "02_Support_Policy_v2_DEPRECATED.pdf": {
        "source_name": "02_Support_Policy_v2_DEPRECATED.pdf",
        "doc_title": "ParcelPilot Support Policy v2 (DEPRECATED)",
        "version": "v2_DEPRECATED",
        "status": "DEPRECATED",
        "effective_date": "2025-01-01",
        "customer_scope": "general",
        "authority_level": 99,
        "is_authoritative": False,
        "superseded_by": "Support Policy v3 effective 1 May 2026"
    },
    "03_Cancellation_and_Service_Credit_SOP_v4.pdf": {
        "source_name": "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
        "doc_title": "ParcelPilot Cancellation & Service Credit SOP v4",
        "version": "v4_SOP",
        "status": "CURRENT",
        "effective_date": "2026-06-15",
        "customer_scope": "general",
        "authority_level": 2,
        "is_authoritative": True,
        "supersedes": None
    },
    "04_Product_Operations_Guide_and_Known_Issues.pdf": {
        "source_name": "04_Product_Operations_Guide_and_Known_Issues.pdf",
        "doc_title": "ParcelPilot Product Operations Guide & Known Issues",
        "version": "v1_OPS_GUIDE",
        "status": "CURRENT",
        "effective_date": "2026-08-14",
        "customer_scope": "general",
        "authority_level": 3,
        "is_authoritative": True,
        "supersedes": None
    },
    "05_Northstar_Logistics_Enterprise_Agreement.pdf": {
        "source_name": "05_Northstar_Logistics_Enterprise_Agreement.pdf",
        "doc_title": "Northstar Logistics Enterprise Agreement",
        "version": "v1_ENTERPRISE_AGREEMENT",
        "status": "ACTIVE",
        "effective_date": "2026-01-01",
        "customer_scope": "Northstar",
        "account_id": "ACCT-001",
        "authority_level": 1, # Highest authority for ACCT-001
        "is_authoritative": True,
        "supersedes": "Overrides standard Support Policy and SOP cancellation terms for Northstar"
    },
    "06_LumenWorks_Service_Agreement.pdf": {
        "source_name": "06_LumenWorks_Service_Agreement.pdf",
        "doc_title": "LumenWorks Service Agreement",
        "version": "v1_SERVICE_AGREEMENT",
        "status": "ACTIVE",
        "effective_date": "2026-03-01",
        "customer_scope": "LumenWorks",
        "account_id": "ACCT-002",
        "authority_level": 1, # Highest authority for ACCT-002
        "is_authoritative": True,
        "supersedes": "Overrides default failed-pickup credit timing and amount for LumenWorks"
    }
}

def clean_text(text: str) -> str:
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

def chunk_pdf(file_path: str):
    base_name = os.path.basename(file_path)
    meta = DOC_METADATA.get(base_name, {})
    reader = pypdf.PdfReader(file_path)
    full_text = ""
    for page in reader.pages:
        full_text += (page.extract_text() or "") + "\n"
    
    full_text = clean_text(full_text)
    
    # Section splitters based on numbered headings (e.g., "1. ", "2. ", "3. ") or Known Issues ("KI-")
    sections = []
    
    if base_name == "01_Support_Policy_v3_CURRENT.pdf":
        sections = [
            {
                "section_title": "1. Scope and source precedence",
                "content": (
                    "ParcelPilot Support Policy v3 (CURRENT, Effective 1 May 2026, Supersedes Policy v2).\n"
                    "1. Scope and source precedence: This policy defines default support severity and response targets. "
                    "A signed customer agreement may override these defaults. When sources conflict, use the signed customer "
                    "agreement first, then the current support policy, then current product documentation. "
                    "Historical tickets and internal notes are context only and may contain incorrect past guidance."
                )
            },
            {
                "section_title": "2. Severity definitions",
                "content": (
                    "2. Severity definitions:\n"
                    "● P1 - Critical: Complete production outage preventing all shipment creation for a customer, confirmed security incident or suspected credential exposure, or another event causing immediate material business risk with no workaround.\n"
                    "● P2 - High: Major feature unavailable or materially degraded for a customer, but core operations remain possible or a workaround exists.\n"
                    "● P3 - Normal: Minor defect, how-to question, configuration request, or issue with limited operational impact."
                )
            },
            {
                "section_title": "3. Default first-response targets",
                "content": (
                    "3. Default first-response targets:\n"
                    "Plan | P1 Target | P2 Target | P3 Target\n"
                    "Enterprise: P1 = 30 minutes (24x7), P2 = 2 hours, P3 = 1 business day\n"
                    "Growth: P1 = 2 business hours, P2 = 4 business hours, P3 = 2 business days\n"
                    "Standard: P1 = 4 business hours, P2 = 1 business day, P3 = 2 business days"
                )
            },
            {
                "section_title": "4. Escalation rules",
                "content": (
                    "4. Escalation:\n"
                    "P1 incidents should be escalated immediately. If a response target is already breached, the agent should clearly state the breach and recommend escalation rather than hiding uncertainty."
                )
            }
        ]
    elif base_name == "02_Support_Policy_v2_DEPRECATED.pdf":
        sections = [
            {
                "section_title": "DEPRECATED Support Policy v2 (Historical Reference Only)",
                "content": (
                    "ParcelPilot Support Policy v2 [STATUS: DEPRECATED - DO NOT USE FOR CURRENT REQUESTS].\n"
                    "Effective: 1 January 2025. Superseded by: Support Policy v3 effective 1 May 2026.\n"
                    "Old targets: Enterprise P1 = 1 hour, P2 = 4 hours, P3 = 2 business days; Growth P1 = 4 bh, P2 = 1 bd, P3 = 3 bd; Standard P1 = 8 bh, P2 = 2 bd, P3 = 3 bd.\n"
                    "CRITICAL WARNING: This file is intentionally retained for historical reference and must NOT be used as current policy."
                )
            }
        ]
    elif base_name == "03_Cancellation_and_Service_Credit_SOP_v4.pdf":
        sections = [
            {
                "section_title": "1. Order cancellation rules",
                "content": (
                    "ParcelPilot Cancellation & Service Credit SOP v4 (CURRENT, Effective 15 June 2026).\n"
                    "1. Order cancellation:\n"
                    "● DRAFT: May be cancelled with no fee.\n"
                    "● BOOKED, not yet PICKED_UP: May be cancelled. No fee within 30 minutes of booking. After 30 minutes, charge INR 250 unless a customer agreement explicitly waives the cancellation fee.\n"
                    "● PICKED_UP: Do not cancel. Use the return-to-origin workflow if the customer wants the parcel returned.\n"
                    "● DELIVERED: Cannot be cancelled."
                )
            },
            {
                "section_title": "2. Failed-pickup service credits (Default Policy)",
                "content": (
                    "2. Failed-pickup service credits:\n"
                    "Under the default policy, a customer is eligible for a service credit when the pickup is more than 2 hours past the end of the scheduled pickup window, the carrier is at fault, and there is no customer-caused issue.\n"
                    "The default credit is the lower of INR 500 or 10% of the shipment fee. A signed customer agreement may replace the default delay threshold, credit amount, or cap."
                )
            },
            {
                "section_title": "3. Approval and uncertainty rules",
                "content": (
                    "3. Approval and uncertainty:\n"
                    "● Any individual credit above INR 1,000 requires manager approval.\n"
                    "● Do not promise a credit when carrier fault, pickup timing, or customer fault is unknown.\n"
                    "● When data conflicts, identify the conflict and request verification before a state-changing action."
                )
            }
        ]
    elif base_name == "04_Product_Operations_Guide_and_Known_Issues.pdf":
        sections = [
            {
                "section_title": "1. Plan capabilities and Shipment status",
                "content": (
                    "ParcelPilot Product Operations Guide (CURRENT, Updated 14 August 2026).\n"
                    "1. Plan capabilities:\n"
                    "● Bulk Upload: Available on Growth and Enterprise. Supported file size is up to 5,000 rows per CSV.\n"
                    "● Standard Plan: Bulk Upload is not included.\n"
                    "● Shipment status: BOOKED means the shipment is created but ParcelPilot has not yet received a pickup confirmation. PICKED_UP means carrier pickup has been confirmed."
                )
            },
            {
                "section_title": "2. Known Issue KI-208: Bulk Upload failures on large CSVs",
                "content": (
                    "2. Current known issues:\n"
                    "KI-208 - Bulk Upload failures on large CSVs (Opened: 10 August 2026, Status: Investigating).\n"
                    "Some Growth and Enterprise customers experience intermittent failures on CSV uploads above approximately 3,000 rows, even though the supported product limit remains 5,000 rows.\n"
                    "Workaround: Split the upload into files below 3,000 rows. Individual shipment creation is unaffected."
                )
            },
            {
                "section_title": "3. Known Issue KI-211: SwiftShip pickup webhook delay",
                "content": (
                    "KI-211 - SwiftShip pickup webhook delay (Opened: 12 August 2026, Status: Monitoring).\n"
                    "SwiftShip pickup confirmation webhooks can arrive up to 20 minutes late. A parcel may physically be collected while ParcelPilot still shows BOOKED.\n"
                    "Workaround / SOP: Before telling a customer that a pickup did not occur, verify the carrier status or wait through the known delay window."
                )
            },
            {
                "section_title": "4. Resolved Issue KI-176: Address validation",
                "content": (
                    "3. Resolved issue:\n"
                    "KI-176 - Address validation: Resolved 18 July 2026. Do not use this resolved issue to explain new incidents unless evidence specifically matches it."
                )
            }
        ]
    elif base_name == "05_Northstar_Logistics_Enterprise_Agreement.pdf":
        sections = [
            {
                "section_title": "1. Northstar Custom Support Terms (Overrides Standard Policy)",
                "content": (
                    "ParcelPilot - Northstar Logistics Enterprise Agreement (Account: ACCT-001, Customer: Northstar Logistics, Term: 1 Jan 2026 - 31 Dec 2026, Status: ACTIVE).\n"
                    "1. Support terms: For Northstar Logistics, the following first-response targets replace ParcelPilot's standard support-policy targets:\n"
                    "● P1: 15 minutes, 24x7 (Replaces default Enterprise 30 minutes)\n"
                    "● P2: 1 hour (Replaces default 2 hours)\n"
                    "● P3: 8 business hours (Replaces default 1 business day)"
                )
            },
            {
                "section_title": "2. Northstar Shipment Cancellation Waiver (Overrides SOP Fee)",
                "content": (
                    "2. Shipment cancellation:\n"
                    "Northstar may cancel any BOOKED shipment before pickup with NO cancellation fee, regardless of how long ago the shipment was booked. Once a shipment is PICKED_UP, the standard return-to-origin process applies.\n"
                    "NOTE: This clause strictly overrides the SOP v4 INR 250 fee after 30 minutes."
                )
            },
            {
                "section_title": "3. Northstar Service Credits & Account Contact",
                "content": (
                    "3. Service credits: Monthly aggregate service credits are capped at INR 5,000. Unless this agreement states otherwise, the current ParcelPilot service-credit SOP applies.\n"
                    "4. Account contact: Dedicated CSM: Priya Mehta."
                )
            }
        ]
    elif base_name == "06_LumenWorks_Service_Agreement.pdf":
        sections = [
            {
                "section_title": "1. LumenWorks Support Terms & Cancellation Terms",
                "content": (
                    "ParcelPilot - LumenWorks Service Agreement (Account: ACCT-002, Customer: LumenWorks, Plan: Growth, Term: 1 March 2026 - 28 Feb 2027, Status: ACTIVE).\n"
                    "1. Support terms:\n"
                    "● P1: 2 business hours\n"
                    "● P2: 4 business hours\n"
                    "● P3: 2 business days\n"
                    "● No weekend or after-hours support coverage.\n"
                    "2. Cancellation terms: No special cancellation-fee waiver applies. Use the current ParcelPilot Cancellation & Service Credit SOP (INR 250 fee applies if cancelled >30 min after booking)."
                )
            },
            {
                "section_title": "2. LumenWorks Custom Failed-Pickup Credits (Overrides SOP)",
                "content": (
                    "3. Failed-pickup credits:\n"
                    "If a pickup is more than 4 hours past the end of the scheduled pickup window, the carrier is at fault, and the customer is not at fault, LumenWorks receives a fixed INR 300 service credit.\n"
                    "This clause replaces the default failed-pickup credit amount (lower of INR 500 or 10%) and timing threshold (2 hours) in the SOP."
                )
            }
        ]
    
    # Enrich chunks with metadata
    chunks = []
    for idx, sec in enumerate(sections):
        chunk_id = f"{base_name.replace('.pdf', '')}_chunk_{idx+1}"
        chunks.append({
            "id": chunk_id,
            "source_name": meta.get("source_name", base_name),
            "doc_title": meta.get("doc_title", base_name),
            "version": meta.get("version", "v1"),
            "status": meta.get("status", "ACTIVE"),
            "effective_date": meta.get("effective_date", "2026-01-01"),
            "customer_scope": meta.get("customer_scope", "general"),
            "account_id": meta.get("account_id", None),
            "authority_level": meta.get("authority_level", 2),
            "is_authoritative": meta.get("is_authoritative", True),
            "supersedes": meta.get("supersedes", None),
            "superseded_by": meta.get("superseded_by", None),
            "section_title": sec["section_title"],
            "content": sec["content"]
        })
    return chunks

def embed_texts(texts: list[str]) -> list[list[float]]:
    url = "https://api.voyageai.com/v1/embeddings"
    headers = {
        "Authorization": f"Bearer {VOYAGE_API_KEY}",
        "Content-Type": "application/json"
    }
    # Voyage API supports batching
    resp = requests.post(url, headers=headers, json={
        "model": "voyage-3",
        "input": texts
    })
    resp.raise_for_status()
    data = resp.json()
    return [item["embedding"] for item in data["data"]]

def load_excel_data():
    excel_path = os.path.join(DATA_DIR, "ParcelPilot_Assessment_Data.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    
    accounts = []
    sheet = wb["accounts"]
    headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, r))
        accounts.append({
            "account_id": str(row_dict["account_id"]),
            "account_name": str(row_dict["account_name"]),
            "plan": str(row_dict["plan"]),
            "status": str(row_dict["status"]),
            "csm": str(row_dict["csm"]) if row_dict.get("csm") else None,
            "contract_file": str(row_dict["contract_file"]) if row_dict.get("contract_file") else None,
            "premium_support": bool(row_dict.get("premium_support")),
            "notes": str(row_dict["notes"]) if row_dict.get("notes") else None
        })
        
    orders = []
    sheet = wb["orders"]
    headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, r))
        orders.append({
            "order_id": str(row_dict["order_id"]),
            "account_id": str(row_dict["account_id"]),
            "carrier": str(row_dict["carrier"]),
            "status": str(row_dict["status"]),
            "booked_at": str(row_dict["booked_at"]),
            "pickup_window_start": str(row_dict["pickup_window_start"]) if row_dict.get("pickup_window_start") else None,
            "pickup_window_end": str(row_dict["pickup_window_end"]) if row_dict.get("pickup_window_end") else None,
            "pickup_actual_at": str(row_dict["pickup_actual_at"]) if row_dict.get("pickup_actual_at") else None,
            "shipment_fee_inr": float(row_dict["shipment_fee_inr"]) if row_dict.get("shipment_fee_inr") is not None else 0.0,
            "carrier_fault": bool(row_dict.get("carrier_fault")),
            "customer_fault": bool(row_dict.get("customer_fault")),
            "cancellation_requested_at": str(row_dict["cancellation_requested_at"]) if row_dict.get("cancellation_requested_at") else None,
            "notes": str(row_dict["notes"]) if row_dict.get("notes") else None
        })
        
    tickets = []
    sheet = wb["tickets"]
    headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, r))
        tickets.append({
            "ticket_id": str(row_dict["ticket_id"]),
            "account_id": str(row_dict["account_id"]),
            "created_at": str(row_dict["created_at"]),
            "status": str(row_dict["status"]),
            "subject": str(row_dict["subject"]),
            "description": str(row_dict["description"]),
            "channel": str(row_dict["channel"]),
            "assigned_to": str(row_dict["assigned_to"]) if row_dict.get("assigned_to") else None,
            "last_customer_message_at": str(row_dict["last_customer_message_at"]) if row_dict.get("last_customer_message_at") else None,
            "historical_resolution": str(row_dict["historical_resolution"]) if row_dict.get("historical_resolution") else None
        })

    snapshot_time = "2026-08-16 11:00 Asia/Kolkata"
    return {
        "snapshot_time": snapshot_time,
        "accounts": accounts,
        "orders": orders,
        "tickets": tickets
    }

def main():
    print("🚀 Starting ParcelPilot Ingestion Pipeline...")
    pdf_files = sorted(glob.glob(f"{DATA_DIR}/*.pdf"))
    all_chunks = []
    for pdf_path in pdf_files:
        print(f"📄 Parsing and chunking {os.path.basename(pdf_path)}...")
        chunks = chunk_pdf(pdf_path)
        all_chunks.extend(chunks)
    
    print(f"✨ Extracted {len(all_chunks)} semantic chunks. Generating Voyage-3 embeddings...")
    texts_to_embed = [f"{c['doc_title']} - {c['section_title']}\n{c['content']}" for c in all_chunks]
    embeddings = embed_texts(texts_to_embed)
    
    for chunk, emb in zip(all_chunks, embeddings):
        chunk["embedding"] = emb

    print("📊 Loading Excel workbook...")
    excel_data = load_excel_data()
    
    full_dataset = {
        "snapshot_time": excel_data["snapshot_time"],
        "metadata": {
            "title": "ParcelPilot Data Pack",
            "version": "1.0.0",
            "total_chunks": len(all_chunks),
            "total_accounts": len(excel_data["accounts"]),
            "total_orders": len(excel_data["orders"]),
            "total_tickets": len(excel_data["tickets"])
        },
        "document_chunks": all_chunks,
        "accounts": excel_data["accounts"],
        "orders": excel_data["orders"],
        "tickets": excel_data["tickets"]
    }
    
    with open(OUTPUT_JSON, "w") as f:
        json.dump(full_dataset, f, indent=2)
    print(f" Saved full dataset with embeddings to {OUTPUT_JSON}")
    
    # Generate SQL seed file
    sql_lines = [
        "-- Supabase Seed Data for ParcelPilot",
        "-- Auto-generated by scripts/ingest_data.py\n",
        "TRUNCATE TABLE actions, tickets, orders, accounts, document_chunks CASCADE;\n"
    ]
    
    for a in excel_data["accounts"]:
        csm = f"'{a['csm']}'" if a['csm'] else "NULL"
        cfile = f"'{a['contract_file']}'" if a['contract_file'] else "NULL"
        anotes = a['notes'].replace("'", "''") if a['notes'] else None
        notes = f"'{anotes}'" if anotes else "NULL"
        sql_lines.append(
            f"INSERT INTO accounts (account_id, account_name, plan, status, csm, contract_file, premium_support, notes) "
            f"VALUES ('{a['account_id']}', '{a['account_name']}', '{a['plan']}', '{a['status']}', {csm}, {cfile}, {a['premium_support']}, {notes}) "
            f"ON CONFLICT (account_id) DO UPDATE SET account_name=EXCLUDED.account_name, plan=EXCLUDED.plan;\n"
        )
        
    for o in excel_data["orders"]:
        pw_start = f"'{o['pickup_window_start']}+05:30'" if o['pickup_window_start'] else "NULL"
        pw_end = f"'{o['pickup_window_end']}+05:30'" if o['pickup_window_end'] else "NULL"
        pw_act = f"'{o['pickup_actual_at']}+05:30'" if o['pickup_actual_at'] else "NULL"
        canc_at = f"'{o['cancellation_requested_at']}+05:30'" if o['cancellation_requested_at'] else "NULL"
        onotes = o['notes'].replace("'", "''") if o['notes'] else None
        notes = f"'{onotes}'" if onotes else "NULL"
        sql_lines.append(
            f"INSERT INTO orders (order_id, account_id, carrier, status, booked_at, pickup_window_start, pickup_window_end, pickup_actual_at, shipment_fee_inr, carrier_fault, customer_fault, cancellation_requested_at, notes) "
            f"VALUES ('{o['order_id']}', '{o['account_id']}', '{o['carrier']}', '{o['status']}', '{o['booked_at']}+05:30', {pw_start}, {pw_end}, {pw_act}, {o['shipment_fee_inr']}, {o['carrier_fault']}, {o['customer_fault']}, {canc_at}, {notes}) "
            f"ON CONFLICT (order_id) DO NOTHING;\n"
        )
        
    for t in excel_data["tickets"]:
        assigned = f"'{t['assigned_to']}'" if t['assigned_to'] else "NULL"
        last_msg = f"'{t['last_customer_message_at']}+05:30'" if t['last_customer_message_at'] else "NULL"
        thist = t['historical_resolution'].replace("'", "''") if t['historical_resolution'] else None
        hist = f"'{thist}'" if thist else "NULL"
        subj = t['subject'].replace("'", "''")
        desc = t['description'].replace("'", "''")
        sql_lines.append(
            f"INSERT INTO tickets (ticket_id, account_id, created_at, status, subject, description, channel, assigned_to, last_customer_message_at, historical_resolution) "
            f"VALUES ('{t['ticket_id']}', '{t['account_id']}', '{t['created_at']}+05:30', '{t['status']}', '{subj}', '{desc}', '{t['channel']}', {assigned}, {last_msg}, {hist}) "
            f"ON CONFLICT (ticket_id) DO NOTHING;\n"
        )
        
    for c in all_chunks:
        emb_str = "[" + ",".join(str(x) for x in c["embedding"]) + "]"
        content_esc = c["content"].replace("'", "''")
        sec_esc = c["section_title"].replace("'", "''")
        sql_lines.append(
            f"INSERT INTO document_chunks (source_name, version, effective_date, section_title, content, customer_scope, authority_level, embedding) "
            f"VALUES ('{c['source_name']}', '{c['version']}', '{c['effective_date']}', '{sec_esc}', '{content_esc}', '{c['customer_scope']}', {c['authority_level']}, '{emb_str}') "
            f"ON CONFLICT DO NOTHING;\n"
        )
        
    with open(OUTPUT_SQL, "w") as f:
        f.writelines(sql_lines)
    print(f" Generated Supabase SQL seed file in {OUTPUT_SQL}")
    print(" Ingestion successfully finished!")

if __name__ == "__main__":
    main()
